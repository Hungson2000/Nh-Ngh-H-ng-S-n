"""
export_excel.py
----------------
Xuất báo cáo doanh thu + danh sách booking ra file Excel.
Chạy: python export_excel.py
File output: bao_cao_nhanghi.xlsx
"""

import os
from datetime import datetime
from dotenv import load_dotenv
from pymongo import MongoClient
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

load_dotenv()

MONGO_URI = os.getenv("MONGO_URI")
DB_NAME = os.getenv("DB_NAME", "nhanghi")


def get_db():
    client = MongoClient(MONGO_URI)
    return client[DB_NAME]


def parse_date(val):
    if not val:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%dT%H:%M:%S.%fZ", "%Y-%m-%d"):
        try:
            return datetime.strptime(str(val)[:10], fmt[:len(fmt)])
        except:
            continue
    return None


def export_excel():
    db = get_db()
    print("✅ Kết nối MongoDB thành công!")

    # ---- Sheet 1: Danh sách booking ----
    bookings = list(db.bookings.find())
    df_booking = pd.DataFrame(bookings)

    if not df_booking.empty:
        cols = ["code", "name", "room", "phone", "checkin", "checkout", "total", "status"]
        cols = [c for c in cols if c in df_booking.columns]
        df_booking = df_booking[cols].copy()
        df_booking["total"] = pd.to_numeric(df_booking.get("total", 0), errors="coerce").fillna(0)
        df_booking.columns = [c.upper() for c in df_booking.columns]

    # ---- Sheet 2: Doanh thu theo tháng ----
    if not df_booking.empty and "TOTAL" in df_booking.columns:
        date_col = "CREATEDAT" if "CREATEDAT" in df_booking.columns else "CHECKIN"
        if date_col in df_booking.columns:
            df_booking["NGAY"] = df_booking[date_col].apply(parse_date)
            df_booking["THANG"] = df_booking["NGAY"].apply(
                lambda x: x.strftime("%m/%Y") if x else None
            )
            df_doanh_thu = df_booking.groupby("THANG")["TOTAL"].agg(
                SO_BOOKING="count",
                TONG_DOANH_THU="sum"
            ).reset_index()
        else:
            df_doanh_thu = pd.DataFrame(columns=["THANG", "SO_BOOKING", "TONG_DOANH_THU"])
    else:
        df_doanh_thu = pd.DataFrame(columns=["THANG", "SO_BOOKING", "TONG_DOANH_THU"])

    # ---- Sheet 3: Đánh giá ----
    reviews = list(db.reviews.find())
    df_review = pd.DataFrame(reviews)
    if not df_review.empty:
        cols_r = ["room", "star", "comment", "date"]
        cols_r = [c for c in cols_r if c in df_review.columns]
        df_review = df_review[cols_r].copy()
        df_review.columns = [c.upper() for c in df_review.columns]

    # ---- Xuất ra Excel ----
    filename = f"bao_cao_nhanghi_{datetime.now().strftime('%d%m%Y_%H%M')}.xlsx"

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        # Sheet 1
        if not df_booking.empty:
            df_booking.drop(columns=["NGAY", "THANG"], errors="ignore").to_excel(
                writer, sheet_name="Danh sách Booking", index=False
            )
        # Sheet 2
        df_doanh_thu.to_excel(writer, sheet_name="Doanh thu theo tháng", index=False)
        # Sheet 3
        if not df_review.empty:
            df_review.to_excel(writer, sheet_name="Đánh giá", index=False)

        # Format header cho tất cả sheet
        wb = writer.book
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(color="FFFFFF", bold=True)

        for ws in wb.worksheets:
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=10)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    print(f"✅ Xuất Excel thành công: {filename}")
    print(f"📁 File lưu tại: {os.path.abspath(filename)}")


if __name__ == "__main__":
    export_excel()
